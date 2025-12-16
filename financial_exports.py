import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from io import BytesIO
from typing import Dict, List
from datetime import datetime
from format_utils import format_currency

class FinancialExporter:
    """Export financial models to Excel format"""
    
    def __init__(self):
        self.header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        self.header_font = Font(bold=True, color='FFFFFF', size=11)
        self.title_font = Font(bold=True, size=14)
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    def export_cashflow_model(
        self,
        project_name: str,
        model_name: str,
        cashflow_data: Dict,
        metrics: Dict,
        assumptions: Dict
    ) -> bytes:
        """
        Export complete cash flow model to Excel
        
        Args:
            project_name: Name of the mining project
            model_name: Name of the financial model
            cashflow_data: Dictionary with years, revenue, costs, etc.
            metrics: NPV, IRR, payback, etc.
            assumptions: Input assumptions (prices, costs, etc.)
        
        Returns:
            Excel file as bytes
        """
        wb = Workbook()
        
        self._add_summary_sheet(wb, project_name, model_name, metrics, assumptions)
        
        self._add_cashflow_sheet(wb, cashflow_data)
        
        self._add_assumptions_sheet(wb, assumptions)
        
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        return buffer.getvalue()
    
    def _add_summary_sheet(self, wb: Workbook, project_name: str, model_name: str, metrics: Dict, assumptions: Dict):
        """Add summary sheet with key metrics"""
        if 'Sheet' in wb.sheetnames:
            ws = wb['Sheet']
            ws.title = 'Summary'
        else:
            ws = wb.create_sheet('Summary', 0)
        
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 20
        
        row = 1
        
        ws.merge_cells(f'A{row}:B{row}')
        cell = ws[f'A{row}']
        cell.value = 'Financial Model Summary'
        cell.font = self.title_font
        row += 2
        
        ws[f'A{row}'] = 'Project Name:'
        ws[f'B{row}'] = project_name
        ws[f'A{row}'].font = Font(bold=True)
        row += 1
        
        ws[f'A{row}'] = 'Model Name:'
        ws[f'B{row}'] = model_name
        ws[f'A{row}'].font = Font(bold=True)
        row += 1
        
        ws[f'A{row}'] = 'Generated:'
        ws[f'B{row}'] = datetime.now().strftime('%Y-%m-%d %H:%M')
        ws[f'A{row}'].font = Font(bold=True)
        row += 2
        
        ws.merge_cells(f'A{row}:B{row}')
        cell = ws[f'A{row}']
        cell.value = 'Key Financial Metrics'
        cell.font = Font(bold=True, size=12)
        cell.fill = PatternFill(start_color='E7E6E6', end_color='E7E6E6', fill_type='solid')
        row += 1
        
        ws[f'A{row}'] = 'Net Present Value (NPV)'
        ws[f'B{row}'] = format_currency(metrics.get('npv', 0), decimals=2)
        row += 1
        
        ws[f'A{row}'] = 'Internal Rate of Return (IRR)'
        ws[f'B{row}'] = f"{metrics.get('irr', 0):.2f}%" if metrics.get('irr') else "N/A"
        row += 1
        
        ws[f'A{row}'] = 'Payback Period'
        ws[f'B{row}'] = f"{metrics.get('payback', 0):.2f} years" if metrics.get('payback') else "Never"
        row += 1
        
        ws[f'A{row}'] = 'Total Production'
        ws[f'B{row}'] = f"{assumptions.get('total_production', 0):,.0f} tonnes"
        row += 2
        
        ws.merge_cells(f'A{row}:B{row}')
        cell = ws[f'A{row}']
        cell.value = 'Base Assumptions'
        cell.font = Font(bold=True, size=12)
        cell.fill = PatternFill(start_color='E7E6E6', end_color='E7E6E6', fill_type='solid')
        row += 1
        
        ws[f'A{row}'] = 'Mine Life'
        ws[f'B{row}'] = f"{assumptions.get('mine_life', 0)} years"
        row += 1
        
        ws[f'A{row}'] = 'Commodity Price'
        ws[f'B{row}'] = f"${assumptions.get('commodity_price', 0):,.2f} per unit"
        row += 1
        
        ws[f'A{row}'] = 'Initial CAPEX'
        ws[f'B{row}'] = format_currency(assumptions.get('initial_capex', 0), decimals=2)
        row += 1
        
        ws[f'A{row}'] = 'OPEX per unit'
        ws[f'B{row}'] = f"${assumptions.get('opex_per_unit', 0):,.2f}"
        row += 1
        
        ws[f'A{row}'] = 'Discount Rate'
        ws[f'B{row}'] = f"{assumptions.get('discount_rate', 0):.1f}%"
        row += 1
    
    def _add_cashflow_sheet(self, wb: Workbook, cashflow_data: Dict):
        """Add detailed cash flow table"""
        ws = wb.create_sheet('Cash Flow')
        
        df = pd.DataFrame({
            'Year': cashflow_data['years'],
            'Production (tonnes)': cashflow_data['production'],
            'Revenue ($M)': cashflow_data['revenue'],
            'Operating Costs ($M)': cashflow_data['operating_costs'],
            'CAPEX ($M)': cashflow_data['capex'],
            'Royalties ($M)': cashflow_data.get('royalties', [0] * len(cashflow_data['years'])),
            'EBITDA ($M)': cashflow_data['ebitda'],
            'Taxes ($M)': cashflow_data['taxes'],
            'Net Cash Flow ($M)': cashflow_data['net_cashflow']
        })
        
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                cell = ws.cell(row=r_idx, column=c_idx)
                cell.value = value
                cell.border = self.border
                cell.alignment = Alignment(horizontal='right' if c_idx > 1 else 'center')
                
                if r_idx == 1:
                    cell.fill = self.header_fill
                    cell.font = self.header_font
                    cell.alignment = Alignment(horizontal='center')
        
        for col in range(1, 10):
            ws.column_dimensions[chr(64 + col)].width = 15
        
        ws.column_dimensions['A'].width = 8
    
    def _add_assumptions_sheet(self, wb: Workbook, assumptions: Dict):
        """Add detailed assumptions sheet"""
        ws = wb.create_sheet('Assumptions')
        
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 20
        
        row = 1
        
        ws[f'A{row}'] = 'Parameter'
        ws[f'B{row}'] = 'Value'
        ws[f'A{row}'].font = self.header_font
        ws[f'B{row}'].font = self.header_font
        ws[f'A{row}'].fill = self.header_fill
        ws[f'B{row}'].fill = self.header_fill
        row += 1
        
        for key, value in assumptions.items():
            ws[f'A{row}'] = key.replace('_', ' ').title()
            if isinstance(value, (int, float)):
                ws[f'B{row}'] = value
            else:
                ws[f'B{row}'] = str(value)
            row += 1
    
    def export_sensitivity_analysis(
        self,
        project_name: str,
        sensitivity_results: List[Dict],
        variable_name: str
    ) -> bytes:
        """
        Export sensitivity analysis results to Excel
        
        Args:
            project_name: Name of the project
            sensitivity_results: List of sensitivity results
            variable_name: Name of the variable being analyzed
        
        Returns:
            Excel file as bytes
        """
        wb = Workbook()
        ws = wb.active
        ws.title = 'Sensitivity Analysis'
        
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
        
        row = 1
        
        ws.merge_cells(f'A{row}:D{row}')
        cell = ws[f'A{row}']
        cell.value = f'Sensitivity Analysis: {variable_name.replace("_", " ").title()}'
        cell.font = self.title_font
        row += 1
        
        ws[f'A{row}'] = f'Project: {project_name}'
        row += 2
        
        headers = ['Change (%)', 'Value', 'NPV ($M)', 'IRR (%)']
        for c_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=c_idx)
            cell.value = header
            cell.font = self.header_font
            cell.fill = self.header_fill
        row += 1
        
        for result in sensitivity_results:
            ws.cell(row=row, column=1).value = f"{result['variation_pct']:+.0f}%"
            ws.cell(row=row, column=2).value = result['value']
            ws.cell(row=row, column=3).value = result['npv']
            ws.cell(row=row, column=4).value = result['irr']
            row += 1
        
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        return buffer.getvalue()


def create_financial_exporter() -> FinancialExporter:
    """Factory function to create FinancialExporter instance"""
    return FinancialExporter()
